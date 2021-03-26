from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#respsents a cell in the workbook
class cell_base:
    def get_row(self):
        raise NotImplementedError
    def get_col(self):
        raise NotImplementedError
    def get_value(self):
        raise NotImplementedError


class cell(cell_base):
    def __init__(self,row,col,value):
        self.row = row
        self.col = col
        self.value = value
    
    def get_row(self):
        return self.row
    def get_col(self):
        return self.col
    def get_value(self):
        return self.value


class row_colum_labeler(cell):
    def __init__(self,inital_rows = {"":0},inital_cols = {"":0}):
        self.columns = inital_cols
        self.rows = inital_rows
        self.column_index = max([self.columns[k] for k in self.columns.keys()]+[0])+1
        self.rows_index = max([self.rows[k] for k in self.rows.keys()]+[0])+1
    def get_rows(self):
        return self.rows
    def get_columns(self):
        return self.columns
    def add_row(self,idenifier):
        try:
            return self.rows[idenifier]
        except KeyError:
            self.rows[idenifier] = self.rows_index
            self.rows_index = self.rows_index + 1
            return self.rows[idenifier]
    def add_column(self,idenifier):
        try:
            return self.columns[idenifier]
        except KeyError:
            self.columns[idenifier] = self.column_index
            self.column_index = self.column_index + 1
            return self.columns[idenifier]
    def add_by_cell(self,cell_type):
        self.add_row(cell_type.get_row())
        self.add_column(cell_type.get_col())


def create_work_book(file_name,cells,index_labeler,yshift=1,xshift=1):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "processes"
    
    for column_title in index_labeler.get_columns().keys():
        ws1.cell(row = xshift, column = index_labeler.get_columns()[column_title]+yshift,value = column_title)
           
    for row_title in index_labeler.get_rows().keys():
        ws1.cell(column = yshift, row = index_labeler.get_rows()[row_title]+xshift,value = row_title)
    

    for cell in cells:
        ws1.cell(row=index_labeler.get_rows()[cell.get_row()]+yshift,column = index_labeler.get_columns()[cell.get_col()]+xshift,value=cell.get_value())

    wb.save(filename = file_name)

if __name__ == "__main__":
    cells = []

    column_titles = {"id":0,"name":1,"number":2}
    row_titles = {"id":0,"id_num1":1,"id_num2":2}
    indexs = row_colum_labeler(row_titles,column_titles)
    cells.append(cell("id_num1","name","alex"))
    cells.append(cell("id_num1","number","971"))
    cells.append(cell("id_num2","name","hubert"))
    cells.append(cell("id_num2","number","503"))


    create_work_book("./test_book.xlsx",cells,indexs)

